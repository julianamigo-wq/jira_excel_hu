import openpyxl
import csv
import uuid
from io import StringIO
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path # Necesitamos Path para manejar rutas de forma limpia

def createxlsx(csv_text: str, target_dir: Path, name_issue: str):
    """
    Convierte el texto CSV en un archivo XLSX, aplica formato y lo guarda
    en el directorio especificado.

    :param csv_text: El string de datos en formato CSV (obtenido de la IA).
    :param target_dir: La ruta (Path object) donde se debe guardar el archivo XLSX.
    """
    # --- 1. CONFIGURACIÓN DE ESTILOS ---
    # ... (Sin cambios) ...
    verde_claro = "CCFFCC" 
    fill_header = PatternFill(start_color=verde_claro, end_color=verde_claro, fill_type="solid")
    font_header = Font(name='Calibri', size=12, bold=True, color="000000")
    alignment_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- 2. PREPARACIÓN Y PARSEO DE DATOS ---
    csv_file = StringIO(csv_text.strip())
    csv_reader = csv.reader(csv_file)

    # 3. CREACIÓN DEL LIBRO Y ESCRITURA
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Casos de Prueba"

    es_encabezado = True 

    # Escribir las filas en la hoja de cálculo
    for row_data in csv_reader:
        if not row_data or all(not cell for cell in row_data):
            continue

        sheet.append(row_data)

        if es_encabezado:
            row_number = sheet.max_row
            for cell in sheet[row_number]:
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = alignment_header
            es_encabezado = False
        
    # --- 4. AUTO-AJUSTE Y GUARDADO ---

    # Auto-ajustar el ancho de las columnas (optimizado para el contenido)
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                cell_length = len(str(cell.value)) * 1.2
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        if max_length > 60:
            max_length = 60
        if max_length < 10:
            max_length = 10
            
        sheet.column_dimensions[column].width = max_length

    # 5. CONSTRUCCIÓN DE LA RUTA DE GUARDADO Y ALMACENAMIENTO (¡NUEVA LÓGICA!)
    
    # Generar un ID único corto (ej: '5a4c3f9e')
    unique_id = uuid.uuid4().hex[:8]
    
    # Construcción del nombre: CP_NAMEISSUE_UNIQUEID.xlsx
    # Ejemplo: CP_PROYECTO-123_5a4c3f9e.xlsx
    nombre_base = f"CP_{name_issue}_{unique_id}.xlsx"
    
    # Combinar la ruta de destino (target_dir) con el nombre del archivo
    ruta_guardado = target_dir / nombre_base 
    
    # Guardar el libro de trabajo en la ruta completa
    workbook.save(ruta_guardado)

    print(f"Archivo '{nombre_base}' creado exitosamente en: {ruta_guardado}")
